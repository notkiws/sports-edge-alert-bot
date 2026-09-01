"""Tennis-Data workbook ingestion for ATP/WTA 500+ singles history."""

import json
from collections.abc import Iterable
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from sports_edge.domain.tennis import (
    HistoricalTennisMatch,
    TennisMatchStatus,
    TennisTour,
    TournamentLevel,
)

LEVEL_LABELS: dict[TennisTour, dict[str, TournamentLevel]] = {
    TennisTour.ATP: {
        "ATP500": TournamentLevel.LEVEL_500,
        "Masters 1000": TournamentLevel.LEVEL_1000,
        "Masters Cup": TournamentLevel.FINALS,
        "Grand Slam": TournamentLevel.GRAND_SLAM,
    },
    TennisTour.WTA: {
        "WTA500": TournamentLevel.LEVEL_500,
        "WTA1000": TournamentLevel.LEVEL_1000,
        "Tour Championships": TournamentLevel.FINALS,
        "Grand Slam": TournamentLevel.GRAND_SLAM,
    },
}


class InvalidTennisDataWorkbook(ValueError):
    """Raised when required Tennis-Data fields are missing or invalid."""


def _rank(value: object) -> int | None:
    if value is None or str(value).strip().upper() in {"", "NR", "N/A"}:
        return None
    try:
        if isinstance(value, bool):
            raise ValueError
        if isinstance(value, int):
            return value
        if isinstance(value, float) and value.is_integer():
            return int(value)
        if isinstance(value, str):
            return int(value.strip())
        raise ValueError
    except (TypeError, ValueError) as error:
        raise InvalidTennisDataWorkbook(f"invalid ranking value {value!r}") from error


def _played_on(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raise InvalidTennisDataWorkbook(f"invalid match date {value!r}")


class TennisDataWorkbookReader:
    """Normalize eligible rows while preserving source-file provenance."""

    def read(self, path: Path, tour: TennisTour) -> tuple[HistoricalTennisMatch, ...]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if not workbook.worksheets:
            raise InvalidTennisDataWorkbook("workbook has no worksheets")
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as error:
            raise InvalidTennisDataWorkbook("workbook is empty") from error
        headers = {str(name): index for index, name in enumerate(raw_headers) if name is not None}
        level_field = next(
            (name for name in ("Series", "Tier", "Level") if name in headers),
            None,
        )
        required = {
            "Location",
            "Tournament",
            "Date",
            "Court",
            "Surface",
            "Round",
            "Winner",
            "Loser",
            "WRank",
            "LRank",
            "Comment",
        }
        missing = required.difference(headers)
        if missing or level_field is None:
            missing_names = missing | ({"Level"} if level_field is None else set())
            raise InvalidTennisDataWorkbook(
                f"missing required columns: {sorted(missing_names)}"
            )

        matches: list[HistoricalTennisMatch] = []
        for source_row, row in enumerate(rows, start=2):
            level_label = str(row[headers[level_field]])
            level = LEVEL_LABELS[tour].get(level_label)
            if level is None:
                continue
            round_name = str(row[headers["Round"]])
            if "qual" in round_name.casefold():
                continue
            try:
                status = TennisMatchStatus(str(row[headers["Comment"]]))
            except ValueError as error:
                raise InvalidTennisDataWorkbook(
                    f"unknown match status at row {source_row}"
                ) from error
            matches.append(
                HistoricalTennisMatch(
                    source="tennis-data.co.uk",
                    source_file=path.name,
                    source_row=source_row,
                    tour=tour,
                    tournament=str(row[headers["Tournament"]]),
                    location=str(row[headers["Location"]]),
                    played_on=_played_on(row[headers["Date"]]),
                    level=level,
                    court=str(row[headers["Court"]]),
                    surface=str(row[headers["Surface"]]),
                    round_name=round_name,
                    winner=str(row[headers["Winner"]]),
                    loser=str(row[headers["Loser"]]),
                    winner_rank=_rank(row[headers["WRank"]]),
                    loser_rank=_rank(row[headers["LRank"]]),
                    status=status,
                )
            )
        return tuple(matches)


def write_matches_jsonl(
    matches: Iterable[HistoricalTennisMatch],
    destination: Path,
) -> Path:
    """Atomically persist normalized matches with source provenance."""
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(destination.suffix + ".tmp")
    with temporary.open("w", encoding="utf-8") as output:
        for match in matches:
            record = {
                "source": match.source,
                "source_file": match.source_file,
                "source_row": match.source_row,
                "tour": match.tour.value,
                "tournament": match.tournament,
                "location": match.location,
                "played_on": match.played_on.isoformat(),
                "level": match.level.value,
                "court": match.court,
                "surface": match.surface,
                "round": match.round_name,
                "winner": match.winner,
                "loser": match.loser,
                "winner_rank": match.winner_rank,
                "loser_rank": match.loser_rank,
                "status": match.status.value,
            }
            output.write(json.dumps(record, sort_keys=True) + "\n")
    temporary.replace(destination)
    return destination
